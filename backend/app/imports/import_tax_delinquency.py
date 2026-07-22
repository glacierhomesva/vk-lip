import argparse
import csv
import re
from collections.abc import Iterable
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from sqlalchemy.orm import Session

from app.db.database import SessionLocal
from app.models.assessment import Assessment  # noqa: F401
from app.models.parcel import Parcel

ROOT = Path(__file__).resolve().parents[3]
DEFAULT_CSV_FILE = ROOT / "data/raw/tax_delinquency.csv"


def empty_or_none(value: Optional[object]) -> Optional[str]:
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped if stripped else None


def float_or_none(value: Optional[object]) -> Optional[float]:
    normalized = empty_or_none(value)
    if normalized is None:
        return None
    try:
        return float(normalized.replace(",", ""))
    except ValueError:
        return None


def _normalize_key(value: object) -> str:
    return "".join(character for character in str(value).strip().lower() if character.isalnum())


def _get_value(row: dict[str, object], *aliases: str) -> object | None:
    normalized = {_normalize_key(key): value for key, value in row.items() if key is not None}
    for alias in aliases:
        value = normalized.get(_normalize_key(alias))
        if value is not None:
            return value
    return None


def _normalize_headers(headers: Iterable[object]) -> list[str]:
    return [_normalize_key(header) for header in headers if empty_or_none(header) is not None]


def _looks_like_header(headers: Iterable[object]) -> bool:
    normalized = set(_normalize_headers(headers))
    if not normalized:
        return False

    parcel_aliases = {
        _normalize_key(alias)
        for alias in (
            "parcel_number",
            "ParcelNumber",
            "Parcel Number",
            "Parcel#",
            "APN",
            "Parcel ID",
            "PIN #",
        )
    }
    amount_aliases = {
        _normalize_key(alias)
        for alias in (
            "lien_amount",
            "LienAmount",
            "Lien Amount",
            "amount",
            "tax_lien_amount",
            "TaxLienAmount",
            "Total",
        )
    }

    return bool(normalized & parcel_aliases) or bool(normalized & amount_aliases)


def _read_csv_rows(path: Path) -> Iterable[dict[str, object]]:
    with path.open(newline="", encoding="utf-8-sig") as stream:
        reader = csv.reader(stream)
        headers: list[str] | None = None

        for row in reader:
            if not any(empty_or_none(cell) for cell in row):
                continue

            if headers is None:
                if not _looks_like_header(row):
                    continue
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                continue

            yield dict(zip(headers, row, strict=False))


def _read_xlsx_rows(path: Path) -> Iterable[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return

        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        for row in rows:
            values = dict(zip(headers, row, strict=False))
            yield values
    finally:
        workbook.close()


def _read_rows(path: str) -> Iterable[dict[str, object]]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        return _read_csv_rows(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_rows(file_path)

    raise ValueError(f"Unsupported delinquency file format: {file_path.suffix}. Use .csv or .xlsx")


def _parcel_number_from_row(row: dict[str, object]) -> Optional[str]:
    return empty_or_none(
        _get_value(
            row,
            "parcel_number",
            "ParcelNumber",
            "Parcel Number",
            "Parcel#",
            "APN",
            "Parcel ID",
            "PIN #",
        ),
    )


def _lien_amount_from_row(row: dict[str, object]) -> Optional[float]:
    return float_or_none(
        _get_value(
            row,
            "lien_amount",
            "LienAmount",
            "Lien Amount",
            "amount",
            "tax_lien_amount",
            "TaxLienAmount",
            "Total",
        ),
    )


def _account_number_from_row(row: dict[str, object]) -> Optional[str]:
    return empty_or_none(
        _get_value(
            row,
            "Account",
            "Account Number",
            "account_number",
        ),
    )


def _read_account_totals_from_detailed_csv(path: Path) -> dict[str, float]:
    account_totals: dict[str, float] = {}
    current_account: str | None = None

    with path.open(newline="", encoding="utf-8-sig") as stream:
        reader = csv.reader(stream)

        for row in reader:
            first_cell = row[0].strip() if row else ""
            if first_cell:
                match = re.search(r"Account:\s*(\d+)", first_cell)
                if match:
                    current_account = match.group(1)
                    continue

            if current_account is None or len(row) <= 38:
                continue

            total_amount = float_or_none(row[38])
            if total_amount is None:
                continue

            account_totals[current_account] = total_amount
            current_account = None

    return account_totals


def _chunked(values: list[str], size: int) -> Iterable[list[str]]:
    for start in range(0, len(values), size):
        yield values[start : start + size]


def import_tax_delinquency_from_files(paths: list[str], db: Session, batch_size: int = 1000) -> tuple[int, int]:
    records: dict[str, Optional[float]] = {}
    account_to_parcel: dict[str, str] = {}
    account_to_total: dict[str, float] = {}
    skipped = 0

    if not paths:
        raise ValueError("Provide at least one delinquency file path.")

    for path in paths:
        file_path = Path(path)
        if not file_path.exists():
            raise FileNotFoundError(
                f"Tax delinquency source not found at {file_path}. Add a CSV or Excel file with parcel_number and optional lien_amount columns.",
            )

        if file_path.suffix.lower() == ".csv":
            account_to_total.update(_read_account_totals_from_detailed_csv(file_path))

        for row in _read_rows(str(file_path)):
            parcel_number = _parcel_number_from_row(row)
            account_number = _account_number_from_row(row)

            if account_number and parcel_number:
                account_to_parcel[account_number] = parcel_number

            if parcel_number is None:
                skipped += 1
                continue

            lien_amount = _lien_amount_from_row(row)
            if lien_amount is not None or parcel_number not in records:
                records[parcel_number] = lien_amount

    for account_number, total_amount in account_to_total.items():
        parcel_number = account_to_parcel.get(account_number)
        if parcel_number is None:
            continue

        records[parcel_number] = total_amount

    # Treat CSV input as a full snapshot and clear stale delinquency flags first.
    db.query(Parcel).filter(Parcel.tax_delinquent.is_(True)).update(
        {
            Parcel.tax_delinquent: False,
            Parcel.tax_lien_amount: None,
        },
        synchronize_session=False,
    )
    db.commit()

    updated = 0
    parcel_numbers = list(records.keys())

    for batch in _chunked(parcel_numbers, batch_size):
        existing = {
            parcel.parcel_number: parcel
            for parcel in db.query(Parcel).filter(Parcel.parcel_number.in_(batch)).all()
        }

        for parcel_number in batch:
            parcel = existing.get(parcel_number)
            if parcel is None:
                skipped += 1
                continue

            parcel.tax_delinquent = True
            parcel.tax_lien_amount = records[parcel_number]
            updated += 1

        db.commit()

    return updated, skipped


def import_tax_delinquency_from_csv(path: str, db: Session) -> tuple[int, int]:
    return import_tax_delinquency_from_files([path], db)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Import tax delinquency data from CSV or Excel into the parcels table.",
    )
    parser.add_argument(
        "paths",
        nargs="*",
        help="One or more delinquency files (.csv, .xlsx, or .xlsm).",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=1000,
        help="Number of parcels to update per database batch (default: 1000).",
    )
    return parser


def main(csv_paths: list[str] | None = None) -> None:
    if csv_paths is None:
        parser = build_parser()
        args = parser.parse_args()
        paths = args.paths or ([str(DEFAULT_CSV_FILE)] if DEFAULT_CSV_FILE.exists() else [])
        if not paths:
            parser.error("Provide at least one delinquency file path.")
        batch_size = args.batch_size
    else:
        paths = csv_paths
        batch_size = 1000

    db = SessionLocal()
    try:
        updated, skipped = import_tax_delinquency_from_files(paths, db, batch_size=batch_size)
        print(f"Updated tax delinquency for {updated} parcels from {len(paths)} file(s)")
        print(f"Skipped {skipped} rows")
    finally:
        db.close()


if __name__ == "__main__":
    main()